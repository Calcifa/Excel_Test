import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string 

workbook = openpyxl.load_workbook('C:\\Users\DTJ\Desktop\工作簿1.xlsx')
worksheet = workbook.active
#worksheet['A7'] = 13
#worksheet['D7'] = 'stop'
#x=7
#y=6
#letter = str(get_column_letter(x))
#number = str(y)
#coordinate = letter + number
#worksheet[coordinate] = 12
for x in range(1, 15 + 1):
	letter = str(get_column_letter(x))
	for y in range(1, 20 + 1):
		number = str(y)
		coordinate = letter + number
		worksheet[coordinate] = x + y
workbook.save('C:\\Users\DTJ\Desktop\工作簿1.xlsx')